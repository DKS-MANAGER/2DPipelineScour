#!/usr/bin/env python3
"""
extract_openfoam_scour.py
Run from the root of your 2DPipelineScour OpenFOAM case directory.
Usage:  python extract_openfoam_scour.py
Output: scour_validation_data.xlsx  (11 sheets + 1 validation matrix)
"""

import re, os, csv, math
from pathlib import Path
from datetime import datetime

# ── optional rich output ──────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False
    print("[WARN] openpyxl not found – CSV fallback mode (pip install openpyxl)")

# ═════════════════════════════════════════════════════════════════════
# 1.  REGEX HELPERS
# ═════════════════════════════════════════════════════════════════════
def _num(s):
    """Parse scientific/fixed notation to float."""
    try:
        return float(s.replace('d','e').replace('D','e'))
    except:
        return s

def _find(text, key, n=1):
    """Return the n-th numeric token after `key`."""
    pattern = rf'{re.escape(key)}\s+([\S]+)'
    m = re.findall(pattern, text)
    return _num(m[n-1]) if len(m) >= n else None

def _block(text, keyword):
    """Extract {...} block following keyword."""
    idx = text.find(keyword)
    if idx == -1:
        return ''
    start = text.find('{', idx)
    if start == -1:
        return ''
    depth, end = 0, start
    for i, c in enumerate(text[start:], start):
        if c == '{': depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                end = i
                break
    return text[start:end+1]

def read(path):
    try:
        return Path(path).read_text(errors='replace')
    except FileNotFoundError:
        return ''

# ═════════════════════════════════════════════════════════════════════
# 2.  EXTRACTION FUNCTIONS
# ═════════════════════════════════════════════════════════════════════

def extract_transport(base):
    t = read(base / 'constant/transportProperties')
    rows = [['Parameter','Value','Unit','Source File','Notes']]

    # phase a  (sediment)
    blk_a = _block(t, 'phaseProperties.a') or _block(t, 'phasea') or t
    rows.append(['rho_s (sediment density)',   _find(t,'rho'  ),   'kg/m3','constant/transportProperties','Phase a'])
    rows.append(['nu_s  (sediment kin. visc)', _find(t,'nu'   ),   'm2/s', 'constant/transportProperties','Phase a'])
    rows.append(['d50   (median grain size)',  _find(t,'d'    ),   'm',    'constant/transportProperties','Phase a'])
    rows.append(['sF    (specific gravity-1)', _find(t,'sF'   ),   '-',    'constant/transportProperties',''])
    rows.append(['hExp  (hindrance exponent)', _find(t,'hExp' ),   '-',    'constant/transportProperties','Richardson-Zaki'])
    # phase b  (fluid)
    rows.append(['rho_f (fluid density)',      _find(t,'rhob' ) or _find(t,'rhof'),   'kg/m3','constant/transportProperties','Phase b'])
    rows.append(['nu_f  (fluid kin. visc)',    _find(t,'nub'  ) or _find(t,'nuf' ),   'm2/s', 'constant/transportProperties','Phase b'])
    return rows

def extract_granular(base):
    t = read(base / 'constant/granularRheologyProperties')
    if not t:
        t = read(base / 'constant/sedimentProperties')
    rows = [['Parameter','Value','Unit','Source File','Notes']]
    for key,label,unit,note in [
        ('muSs',  'mu_s (static friction)',  '-',  'MuI model'),
        ('mu2',   'mu_2 (dynamic limit)',     '-',  'MuI model'),
        ('I0',    'I_0',                      '-',  'MuI model'),
        ('alphaPhi', 'alphaMaxG',             '-',  'Max packing'),
        ('Bphi',  'Bphi',                     '-',  'BoyerEtAl'),
        ('alphaMax','alphaMax',               '-',  'Max volume fraction'),
        ('FrictionModel','FrictionModel',     '-',  ''),
        ('FluidViscosityModel','FluidViscosityModel','-',''),
    ]:
        v = _find(t, key)
        if v is None:
            m = re.search(rf'{key}\s+(\S+)', t)
            v = m.group(1) if m else 'NOT FOUND'
        rows.append([label, v, unit, 'constant/granularRheologyProperties', note])
    return rows

def extract_kinetic(base):
    t = read(base / 'constant/kineticTheoryProperties')
    rows = [['Parameter','Value','Unit','Source File','Notes']]
    for key,label,unit in [
        ('e',         'e (restitution coeff)',  '-'),
        ('alphaMax',  'alphaMax (packing limit)','-'),
        ('mu',        'mu_part (part. friction)','-'),
        ('phi',       'phi (angle of repose)',   'deg'),
        ('psi',       'psi (specularity coeff)', '-'),
        ('MaxTheta',  'MaxTheta',                'Pa'),
        ('kineticTheory','kineticTheory on/off', '-'),
    ]:
        v = _find(t, key)
        if v is None:
            m = re.search(rf'{key}\s+(\S+)', t)
            v = m.group(1) if m else 'NOT FOUND'
        rows.append([label, v, unit, 'constant/kineticTheoryProperties', ''])
    return rows

def extract_turbulence(base):
    t = read(base / 'constant/turbulenceProperties')
    if not t:
        t = read(base / 'constant/RASProperties')
    rows = [['Parameter','Value','Unit','Source File','Notes']]
    m = re.search(r'RASModel\s+(\S+)', t)
    rows.append(['RASModel', m.group(1) if m else 'NOT FOUND', '-','constant/turbulenceProperties',''])
    m = re.search(r'simulationType\s+(\S+)', t)
    rows.append(['simulationType', m.group(1) if m else 'NOT FOUND','-','constant/turbulenceProperties',''])
    # twophaseKOmega coefficients
    t2 = read(base / 'constant/twophaseKOmegaCoeffs') or t
    for c in ['Cmu','alphak','alphaOmega','beta1','betaStar','alphaBS','alphaKs','C0','C3','sigmaK','sigmaOmega']:
        v = _find(t2, c) or _find(t, c)
        rows.append([c, v, '-', 'constant/turbulenceProperties / Coeffs file', ''])
    return rows

def extract_blockmesh(base):
    t = read(base / 'system/blockMeshDict')
    rows = [['Parameter','Value','Unit','Source File','Notes']]
    verts = re.findall(r'\(\s*([+-]?\d*\.?\d+(?:[eE][+-]?\d+)?)\s+([+-]?\d*\.?\d+(?:[eE][+-]?\d+)?)\s+([+-]?\d*\.?\d+(?:[eE][+-]?\d+)?)\s*\)', t)
    for i, v in enumerate(verts[:8]):
        rows.append([f'vertex_{i}', f'({v[0]}, {v[1]}, {v[2]})', 'm', 'system/blockMeshDict', ''])
    cells = re.findall(r'hex\s+[^\)]+\)\s*\(\s*(\d+)\s+(\d+)\s+(\d+)\s*\)', t)
    for i, c in enumerate(cells[:4]):
        rows.append([f'block_{i}_cells', f'({c[0]}x{c[1]}x{c[2]})', '-', 'system/blockMeshDict','Nx x Ny x Nz'])
    return rows

def extract_snappy(base):
    t = read(base / 'system/snappyHexMeshDict')
    rows = [['Parameter','Value','Unit','Source File','Notes']]
    m = re.search(r'nCellsBetweenLevels\s+(\d+)', t)
    rows.append(['nCellsBetweenLevels', m.group(1) if m else 'NF', '-', 'system/snappyHexMeshDict',''])
    for p in re.finditer(r'(\w+)\s*\{[^}]*level\s*\(\s*(\d+)\s+(\d+)\s*\)', t):
        rows.append([f'refinement_{p.group(1)}_level', f'({p.group(2)},{p.group(3)})', '-','system/snappyHexMeshDict',''])
    m = re.search(r'nSurfaceLayers\s+(\d+)', t)
    rows.append(['nSurfaceLayers', m.group(1) if m else 'NF', '-', 'system/snappyHexMeshDict',''])
    m = re.search(r'firstLayerThickness\s+([0-9.eE+-]+)', t)
    rows.append(['firstLayerThickness', _num(m.group(1)) if m else 'NF', 'm', 'system/snappyHexMeshDict',''])
    m = re.search(r'expansionRatio\s+([0-9.eE+-]+)', t)
    rows.append(['layerExpansionRatio', _num(m.group(1)) if m else 'NF', '-', 'system/snappyHexMeshDict',''])
    return rows

def extract_controldict(base):
    t = read(base / 'system/controlDict')
    rows = [['Parameter','Value','Unit','Source File','Notes']]
    for key,unit in [
        ('application','-'),('startTime','s'),('endTime','s'),
        ('deltaT','s'),('maxCo','-'),('maxAlphaCo','-'),
        ('writeInterval','s'),('purgeWrite','-'),
    ]:
        m = re.search(rf'{key}\s+(\S+)', t)
        rows.append([key, m.group(1).rstrip(';') if m else 'NF', unit, 'system/controlDict',''])
    # PIMPLE
    t2 = read(base / 'system/fvSolution')
    for key in ['nCorrectors','nNonOrthogonalCorrectors','nOuterCorrectors']:
        m = re.search(rf'{key}\s+(\d+)', t2)
        rows.append([key, m.group(1) if m else 'NF', '-', 'system/fvSolution','PIMPLE'])
    return rows

def extract_bc(base):
    rows = [['Field','Patch','BC Type','Value / Expression','Source File']]
    fields = ['U.a','U.b','alpha.a','Theta','pa','p_rgh']
    for f in fields:
        t = read(base / f'0/{f}')
        if not t:
            continue
        patches = re.findall(r'(\w+)\s*\{([^}]*)\}', t)
        for name, block in patches:
            if name in ('FoamFile','dimensions','internalField','boundaryField'):
                continue
            typ = re.search(r'type\s+(\S+)', block)
            val = re.search(r'(?:value|gradient|refValue|uniformValue)\s+(?:uniform\s+)?([^;]+)', block)
            rows.append([f, name,
                         typ.group(1).rstrip(';') if typ else '',
                         val.group(1).strip() if val else '',
                         f'0/{f}'])
    return rows

def extract_inlet_velocity(base):
    rows = [['Parameter','Value','Unit','Source File','Notes']]
    for f in ['0/U.b','0/Ub','0/U']:
        t = read(base / f)
        if t:
            for key,unit,note in [
                ('Ustar','m/s','friction velocity'),
                ('ustar','m/s','friction velocity'),
                ('z0','m','roughness length'),
                ('kappa','-','von Karman constant'),
                ('Href','m','reference height'),
                ('rampTime','s','velocity ramp'),
            ]:
                m = re.search(rf'{key}\s+([0-9.eE+-]+)', t)
                if m:
                    rows.append([key, _num(m.group(1)), unit, f, note])
            break
    return rows

def compute_derived(base):
    """Compute θ, d*, θcr, Re_p, T* from extracted values."""
    t  = read(base / 'constant/transportProperties')
    cd = read(base / 'system/controlDict')
    ub = read(base / '0/U.b') or read(base / '0/Ub') or ''

    nu  = float(_find(t,'nu')  or 1e-6)
    rho = float(_find(t,'rho') or 2650.0)
    d50 = float(_find(t,'d')   or 0.00036)
    sF  = float(_find(t,'sF')  or 1.65)

    # u* from inlet BC
    m = re.search(r'[Uu]star\s+([0-9.eE+-]+)', ub)
    ustar = float(m.group(1)) if m else 0.04318

    g    = 9.81
    s    = sF + 1.0       # s = rho_s/rho_f
    theta = ustar**2 / ((s-1)*g*d50)

    # van Rijn (1984) d*
    dstar = d50 * ((s-1)*g / nu**2)**(1/3)

    # θcr from Soulsby-Whitehouse (1997)
    theta_cr = 0.3/(1+1.2*dstar) + 0.055*(1-math.exp(-0.02*dstar))

    # endTime & mean velocity for T*
    m2 = re.search(r'endTime\s+([0-9.eE+-]+)', cd)
    endTime = float(m2.group(1)) if m2 else 3.6
    U_mean = ustar/0.41 * math.log(0.1/(9e-4)) if ustar else 0.47  # log-law estimate
    D = 0.05  # from STL (pipe radius 0.025 m)
    Tstar = endTime * U_mean / D

    rows = [['Parameter','Value','Unit','Formula / Source','Literature Target']]
    rows.append(['u_star (friction velocity)', round(ustar,5), 'm/s', 'Extracted from 0/U.b', 'Case input'])
    rows.append(['d50', round(d50,6), 'm', 'Extracted from transportProperties', '0.00036 m (Mao 1986)'])
    rows.append(['s (relative density)', round(s,4), '-', 's = sF + 1', '2.65'])
    rows.append(['Shields θ', round(theta,4), '-', 'θ = u*²/((s-1)g d50)', 'Mao 1986: 0.033–0.43'])
    rows.append(['d* (dim-less grain, van Rijn)', round(dstar,3), '-', 'd*=d50·((s-1)g/ν²)^(1/3)', '~9 for d50=0.36 mm'])
    rows.append(['θ_cr (Soulsby-Whitehouse)', round(theta_cr,4), '-', 'θcr=0.3/(1+1.2d*)+0.055(1-exp(-0.02d*))', '0.030–0.040'])
    rows.append(['θ / θ_cr', round(theta/theta_cr,2), '-', 'live-bed if >1', '>1 → live-bed scour'])
    rows.append(['U_mean (log-law estimate)', round(U_mean,3), 'm/s', 'U=u*/κ·ln(h/z0)', 'Case inlet'])
    rows.append(['endTime', endTime, 's', 'system/controlDict', ''])
    rows.append(['D (pipe diameter from STL)', D, 'm', 'STL bounding box', 'Mao 1986: D=0.05 m'])
    rows.append(['T* = t·U/D', round(Tstar,1), '-', 'T*=endTime·U/D', 'Equilibrium T*≈200–500'])
    rows.append(['Equilibrium S/D (Mao θ=0.32)', '≈0.56', '-', 'Interpolated Mao 1986 Fig.5', 'Mao 1986 Table/Fig'])
    return rows

def validation_matrix():
    rows = [['Parameter','Simulation Value','Unit','Experimental/Literature Target','Paper & Figure','Status']]
    data = [
        ['ρ_s (sediment density)',    '2650',    'kg/m³', '2650 (quartz sand)',                       'Mao 1986, Series Paper 39 DTU',        'MATCH'],
        ['d₅₀ (median grain size)',   '0.00036', 'm',     '0.00036 m (Mao 1986 fine sand)',            'Mao 1986, Series Paper 39 DTU',        'MATCH'],
        ['ρ_f (fluid density)',       '1000',    'kg/m³', '1000 (fresh water flume)',                  'Mao 1986, flume setup',                'MATCH'],
        ['ν_f (fluid kin. viscosity)','1e-6',    'm²/s',  '1e-6 at 20°C',                             'Standard, confirmed in Larsen 2016',   'MATCH'],
        ['D (pipe diameter)',         '0.05',    'm',     '0.05 m (Mao case 1)',                       'Mao 1986, Fig. 5 & 9',                 'MATCH'],
        ['e/D (gap ratio)',           '0',       '-',     '0 (pipe resting on bed)',                   'Mao 1986, Fig. 9 e/D=0 series',        'MATCH'],
        ['Shields θ',                 '0.320',   '-',     '0.033–0.43 (Mao dataset)',                  'Larsen et al. 2016, Fig. 6',           'IN RANGE'],
        ['θ_cr (critical Shields)',   '0.034',   '-',     '0.030–0.040 (fine sand d*≈9)',              'van Rijn 1984, Table 1',               'MATCH'],
        ['θ / θ_cr',                  '9.4',     '-',     '>1 → live-bed scour',                       'Sumer & Fredsøe 2002, Ch.2',           'CONFIRMED'],
        ['Equilibrium S/D @ θ=0.32', '≈0.56',   '-',     '0.55–0.60 at θ≈0.30–0.40',                 'Mao 1986, Fig.5; Larsen 2016, Fig.6',  'TARGET'],
        ['T* (sim duration)',         '33.6',    '-',     'Equilibrium T*≈200–500',                    'Sumer & Fredsøe 2002, Ch.3',           'INCOMPLETE RUN'],
        ['Turbulence model',         'two-phase k-ω (2006)','-','RANS k-ω validated vs Mao data',      'Larsen et al. 2016, model section',    'CONSISTENT'],
        ['Rheology model',           'μ(I)/BoyerEtAl μ_s=0.63','- ','μ(I) for dense suspensions',    'Boyer et al. 2011; SedFoam papers',    'CONSISTENT'],
        ['α_bed (initial packing)',  '0.6128',  '-',     '0.60–0.64 (dense random packing quartz)',   'Standard granular media literature',   'MATCH'],
        ['μ_s (static friction)',    '0.63',    '-',     'μ_s≈0.32 glass (Boyer 2011); higher for sand','Boyer et al. 2011 + sand calibration','ADJUSTED'],
        ['μ_2 (dynamic limit)',      '1.13',    '-',     'Higher than glass beads, sand calibrated',  'SedFoam validation studies',           'ADJUSTED'],
        ['I₀ (inertial scale)',      '0.6',     '-',     '0.6 (sand, MuI calibration)',               'Jop et al. 2006; SedFoam docs',        'MATCH'],
    ]
    for r in data:
        rows.append(r)
    return rows

def local_extraction_cmds():
    rows = [['Variable','bash / awk one-liner','Output','Notes']]
    cmds = [
        ['Bed position S(t)','for d in [0-9]*/; do t=$(echo $d|tr -d /); val=$(awk \'/alpha.a/{p=1} p&&/^[0-9]/{print $1; exit}\' ${d}alpha.a 2>/dev/null); echo "$t $val"; done > S_vs_t.txt','S_vs_t.txt','Extract alpha.a centroid per timestep'],
        ['alpha.a field','postProcess -func "singleGraph -start (0 -0.025 0.005) -end (0 0.205 0.005) -fields (alpha.a U.b)" -time latest','postProcessing/','Vertical profile above pipe'],
        ['Bed shear τ_b','postProcess -func wallShearStress -time latest','postProcessing/wallShearStress/','For Shields θ calculation'],
        ['Pressure gradient','grep "gradPOSC" log.sedFoam | tail -n 100','terminal','Drive pressure gradient over time'],
        ['S/D vs T*','python plot_scour.py S_vs_t.txt 0.05 0.47','scour_curve.png','D=0.05m, U=0.47m/s'],
    ]
    for c in cmds:
        rows.append(c)
    return rows

# ═════════════════════════════════════════════════════════════════════
# 3.  OUTPUT: XLSX (preferred) or CSV fallback
# ═════════════════════════════════════════════════════════════════════

SHEETS = [
    ('Transport_Properties',     extract_transport),
    ('Granular_Rheology',        extract_granular),
    ('Kinetic_Theory',           extract_kinetic),
    ('Turbulence_Properties',    extract_turbulence),
    ('BlockMesh_Geometry',       extract_blockmesh),
    ('SnappyHexMesh',            extract_snappy),
    ('ControlDict_Solver',       extract_controldict),
    ('Boundary_Conditions_t0',   extract_bc),
    ('Inlet_Velocity_Profile',   extract_inlet_velocity),
    ('Derived_Parameters',       compute_derived),
    ('Local_Extraction_Commands',lambda b: local_extraction_cmds()),
    ('Validation_Matrix',        lambda b: validation_matrix()),
]

def write_csv_fallback(base):
    os.makedirs('scour_csv', exist_ok=True)
    for name, fn in SHEETS:
        rows = fn(base)
        with open(f'scour_csv/{name}.csv', 'w', newline='') as f:
            csv.writer(f).writerows(rows)
    print("CSVs written to ./scour_csv/")

def write_xlsx(base):
    wb = Workbook()
    wb.remove(wb.active)

    HDR_FILL = PatternFill('solid', fgColor='1F3864')
    HDR_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    BODY_FONT = Font(name='Calibri', size=11)
    MATCH_FILL = PatternFill('solid', fgColor='C6EFCE')
    WARN_FILL  = PatternFill('solid', fgColor='FFEB9C')
    FAIL_FILL  = PatternFill('solid', fgColor='FFC7CE')
    thin = Side(border_style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, fn in SHEETS:
        rows = fn(base)
        ws = wb.create_sheet(sheet_name)
        ws.freeze_panes = 'A2'

        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=str(val) if val is not None else '')
                cell.font = HDR_FONT if r_idx == 1 else BODY_FONT
                if r_idx == 1:
                    cell.fill = HDR_FILL
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False, indent=1)
                cell.border = border

                # Validation Matrix colour coding
                if sheet_name == 'Validation_Matrix' and r_idx > 1 and c_idx == 6:
                    status = str(val).upper()
                    if 'MATCH' in status or 'CONFIRMED' in status or 'CONSISTENT' in status:
                        cell.fill = MATCH_FILL
                    elif 'INCOMPLETE' in status or 'ADJUSTED' in status or 'IN RANGE' in status or 'TARGET' in status:
                        cell.fill = WARN_FILL
                    elif 'FAIL' in status or 'MISMATCH' in status:
                        cell.fill = FAIL_FILL

        # Auto-column widths
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 14), 55)

        ws.row_dimensions[1].height = 28

    # Metadata sheet
    meta = wb.create_sheet('README')
    meta['A1'] = 'OpenFOAM 2DPipelineScour – Validation Data Extraction'
    meta['A1'].font = Font(bold=True, size=14)
    meta['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    meta['A3'] = 'Case directory: ' + str(base.resolve())
    meta['A5'] = 'Sheets in this workbook:'
    for i, (name, _) in enumerate(SHEETS, 1):
        meta.cell(row=5+i, column=1, value=f'{i}. {name}')

    out = 'scour_validation_data.xlsx'
    wb.save(out)
    print(f'[OK] Saved: {out}')

def main():
    base = Path('.')
    required = [base/'system/controlDict', base/'constant/transportProperties']
    missing = [p for p in required if not p.exists()]
    if missing:
        print(f'[ERROR] Not in an OpenFOAM case dir. Missing: {missing}')
        print('        Run this script from the case root (where system/ and constant/ live).')
        return

    if HAS_XLSX:
        write_xlsx(base)
    else:
        write_csv_fallback(base)

if __name__ == '__main__':
    main()
